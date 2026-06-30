import os
import time
import zipfile
import traceback
import asyncio
from io import BytesIO
from typing import Optional
from rich.progress import track
from lxml import etree
import openpyxl
from openpyxl.utils import coordinate_to_tuple
from openpyxl.drawing.spreadsheet_drawing import (
    OneCellAnchor,
    TwoCellAnchor,
)
from .maths.tools import (
    extract_math_from_paragraph as _extract_math_from_paragraph,
    paragraph_has_math as _paragraph_has_math,
)
from ..base import BaseConverter
from ...utils.utils import clean_text_linebreak


# DrawingML main namespace
_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
# Markup Compatibility namespace
_MC = "{http://schemas.openxmlformats.org/markup-compatibility/2006}"
# OPC relationship type for drawings
_REL_DRAWING = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
)
_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _extract_drawing_math(zf, sheet_index: int) -> list:
    """Extract LaTeX formulas from drawing layer of an xlsx sheet.

    Args:
        zf: An already-opened zipfile.ZipFile object.
        sheet_index: Zero-based sheet index.
    """
    results = []
    rels_path = f"xl/worksheets/_rels/sheet{sheet_index + 1}.xml.rels"

    # Read rels (file may not exist if sheet has no drawing)
    try:
        rels_data = zf.read(rels_path)
    except KeyError:
        return results

    # Find drawing relationship targets
    rels_root = etree.fromstring(rels_data)
    drawing_targets = []
    for rel in rels_root.findall(f"{{{_REL_NS}}}Relationship"):
        if rel.get("Type") == _REL_DRAWING:
            target = rel.get("Target", "")
            # "../drawings/drawingX.xml" → "xl/drawings/drawingX.xml"
            if target.startswith("../"):
                target = "xl/" + target[3:]
            elif not target.startswith("xl/"):
                target = "xl/worksheets/" + target
            drawing_targets.append(target)

    for drawing_path in drawing_targets:
        try:
            drawing_data = zf.read(drawing_path)
            drawing_root = etree.fromstring(drawing_data)
        except Exception:
            continue  # silently skip corrupted or missing drawing

        # Iterate over a:p paragraphs under mc:AlternateContent/mc:Choice
        for alt in drawing_root.iter(f"{_MC}AlternateContent"):
            choice = alt.find(f"{_MC}Choice")
            if choice is None:
                continue
            for para in choice.iter(f"{_A}p"):
                if _paragraph_has_math(para):
                    results.extend(_extract_math_from_paragraph(para))

    return results


def _get_sheet_width_emu(ws, openpyxl_mod) -> int:
    """Return the total column width of the worksheet in EMU. 1 char width ~ 7px, 1px = 9525 EMU."""
    CHAR_TO_EMU = 7 * 9525
    total = 0
    for col_idx in range(1, (ws.max_column or 1) + 1):
        col_letter = openpyxl_mod.utils.get_column_letter(col_idx)
        col_dim = ws.column_dimensions.get(col_letter)
        if col_dim and col_dim.width is not None:
            total += col_dim.width * CHAR_TO_EMU
        else:
            total += 8.43 * CHAR_TO_EMU  # default column width
    return int(total)


def _get_image_cx(anchor) -> Optional[int]:
    """Return image display width in EMU for OneCellAnchor; return None for TwoCellAnchor."""
    try:
        if isinstance(anchor, OneCellAnchor):
            return anchor.ext.cx
    except (AttributeError, ImportError):
        pass
    return None


def _find_data_bounds(ws, image_map, max_rows=None):
    """Return (min_row, max_row, min_col, max_col) 1-based, or None if the sheet is empty."""
    min_r = min_c = float("inf")
    max_r = max_c = 0

    # Non-empty cells
    for cell in ws._cells.values():
        if cell.value is not None:
            r, c = cell.row, cell.column
            if max_rows is not None and r > max_rows:
                continue
            min_r = min(min_r, r)
            max_r = max(max_r, r)
            min_c = min(min_c, c)
            max_c = max(max_c, c)

    # Merged cell ranges
    for mr in ws.merged_cells.ranges:
        r1, r2 = mr.min_row, mr.max_row
        if max_rows is not None:
            r2 = min(r2, max_rows)
        if r1 > r2:
            continue
        min_r = min(min_r, r1)
        max_r = max(max_r, r2)
        min_c = min(min_c, mr.min_col)
        max_c = max(max_c, mr.max_col)

    # Image anchors (0-based -> 1-based)
    for img_r0, img_c0 in image_map:
        r, c = img_r0 + 1, img_c0 + 1
        if max_rows is not None and r > max_rows:
            continue
        min_r = min(min_r, r)
        max_r = max(max_r, r)
        min_c = min(min_c, c)
        max_c = max(max_c, c)

    if max_r == 0:
        return None
    return (int(min_r), int(max_r), int(min_c), int(max_c))


class XlsxConverter(BaseConverter):
    """
    Xlsx 转化器
    """
    
    supported_extensions = ["xlsx"]
    supported_mimetypes = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ]
    def __init__(self, **kwargs):
        self.kwargs = kwargs
        
    async def ainvoke_item(
        self,
        file: str | os.PathLike,
        excel_max_rows: Optional[int] = None,
        excel_extract_images: bool = False,
        verbose: bool = True,
        pdflconverter = None,
        **kwargs
    ) -> list:
        res = await asyncio.to_thread(
            self.invoke_item,
            file,
            excel_max_rows,
            excel_extract_images,
            verbose,
            pdflconverter,
            **kwargs
        )
        return res
        
    def invoke_item(
        self,
        file: str | os.PathLike,
        excel_max_rows: Optional[int] = None,
        excel_extract_images: bool = False,
        verbose: bool = True,
        pdflconverter = None,
        **kwargs
    ) -> list:
        res = []
        try:
            parse_callback = kwargs.get("parse_callback")
            parse_callback.on_started(**{
                "file": file
            })
            wb = openpyxl.load_workbook(str(file), read_only=False, data_only=True)
            target_sheets = wb.sheetnames
            total_range = list(enumerate(target_sheets))
            total_range = track(
                total_range, 
                total = len(total_range), 
                description=f"File[{file.name}]:",
                disable=not verbose,
                refresh_per_second=1
            )            
            with zipfile.ZipFile(str(file), "r") as _zf:
                for idx,sheet_name in total_range:
                    if sheet_name not in wb.sheetnames:
                        continue
                    start_time = time.perf_counter()
                    sheet_images: dict = {}
                    sheet_image_counter = 0
                    ws = wb[sheet_name]
                    sheet_idx = list(wb.sheetnames).index(sheet_name)
                    # Total sheet width in EMU, used for image percentage calculation
                    sheet_width_emu = _get_sheet_width_emu(ws, openpyxl)

                    # Floating image map: (0-based row, 0-based col) -> [Image, ...]
                    image_map: dict = {}
                    for img in getattr(ws, "_images", []):
                        anchor = img.anchor
                        if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
                            r, c = anchor._from.row, anchor._from.col
                            image_map.setdefault((r, c), []).append(img)
                        elif isinstance(anchor, str):
                            try:
                                r, c = coordinate_to_tuple(anchor)
                                image_map.setdefault((r - 1, c - 1), []).append(img)
                            except Exception:
                                pass

                    # Merged cell map: (row, col) -> MergedCellRange
                    merge_map = {}
                    for mr in ws.merged_cells.ranges:
                        for r in range(mr.min_row, mr.max_row + 1):
                            for c in range(mr.min_col, mr.max_col + 1):
                                merge_map[(r, c)] = mr

                    # Trim surrounding empty rows/columns
                    bounds = _find_data_bounds(ws, image_map, excel_max_rows)
                    if bounds is None:
                        continue
                    data_min_row, data_max_row, data_min_col, data_max_col = bounds

                    html_parts = ["<table>"]
                    for row_idx in range(data_min_row, data_max_row + 1):
                        html_parts.append("<tr>")
                        for col_idx in range(data_min_col, data_max_col + 1):
                            cell = ws.cell(row_idx, col_idx)
                            mr = merge_map.get((row_idx, col_idx))
                            # Skip non-origin cells in a merged range
                            if mr and (row_idx, col_idx) != (mr.min_row, mr.min_col):
                                continue
                            tag = "th" if row_idx == data_min_row else "td"
                            attrs = ""
                            if mr:
                                cs = (
                                    min(mr.max_col, data_max_col)
                                    - max(mr.min_col, data_min_col)
                                    + 1
                                )
                                rs = (
                                    min(mr.max_row, data_max_row)
                                    - max(mr.min_row, data_min_row)
                                    + 1
                                )
                                if cs > 1:
                                    attrs += f' colspan="{cs}"'
                                if rs > 1:
                                    attrs += f' rowspan="{rs}"'

                            # Cell text
                            value = cell.value
                            text = str(value) if value is not None else ""
                            # Cell-level font formatting (bold/italic/underline/strikethrough)
                            if text:
                                try:
                                    font = cell.font
                                    if font.bold:
                                        text = f"<b>{text}</b>"
                                    if font.italic:
                                        text = f"<i>{text}</i>"
                                    if font.underline:
                                        text = f"<u>{text}</u>"
                                    if font.strike:
                                        text = f"<del>{text}</del>"
                                    vert_align = font.vertAlign
                                    if vert_align == "superscript":
                                        text = f"<sup>{text}</sup>"
                                    elif vert_align == "subscript":
                                        text = f"<sub>{text}</sub>"
                                except Exception:
                                    pass
                            # Hyperlink wrapping
                            if text:
                                try:
                                    hl = cell.hyperlink
                                    if hl and hl.target:
                                        text = f'<a href="{hl.target}">{text}</a>'
                                except Exception:
                                    pass

                            # Floating images
                            cell_images = image_map.get((row_idx - 1, col_idx - 1), [])
                            img_html = []
                            not_ocr = []
                            need_ocr = []
                            for idd,img_obj in enumerate(cell_images):
                                sheet_image_counter += 1
                                ext = (img_obj.format or "png").lower()
                                filename = f"image_{sheet_image_counter}.{ext}"
                                rel_path = f"images/{filename}"
                                try:
                                    ref = img_obj.ref
                                    if isinstance(ref, BytesIO):
                                        ref.seek(0)
                                        img_bytes = ref.read()
                                    else:
                                        img_bytes = img_obj._data()
                                    sheet_images[rel_path] = img_bytes
                                    if not excel_extract_images:
                                        cx_emu = _get_image_cx(img_obj.anchor)
                                        if cx_emu and sheet_width_emu:
                                            pct = min(
                                                round(cx_emu / sheet_width_emu * 100), 100
                                            )
                                            img_html_content = f'<img src="images/{filename}" width="{pct}%">'
                                        else:
                                            img_html_content = f'<img src="images/{filename}">'
                                        not_ocr.append(
                                            [idd, img_html_content]
                                        )
                                            
                                    else:
                                        ### 需要解析
                                        # rawimage = Image.open(BytesIO(img_bytes))
                                        rawimage = img_bytes
                                        need_ocr.append(
                                            [idd, rawimage]
                                        )
                                except:
                                    pass
                            need_ocr_res = []
                            if need_ocr:
                                image_res = pdflconverter.invoke_image(
                                    file = [x[1] for x in need_ocr],
                                    **kwargs
                                )                     
                                for (idd,rawimage), image_item in zip(need_ocr, image_res): 
                                    img_html_content = ''      
                                    if image_item and isinstance(image_item[0], dict):
                                        img_html_content = image_item[0].get('content','') 
                                    need_ocr_res.append(
                                        [idd, img_html_content]
                                    )
                            need_ocr_res = not_ocr + need_ocr_res   
                            need_ocr_res = list(sorted(need_ocr_res, key=lambda x:x[0], reverse=False))
                            for item in need_ocr_res:
                                img_html.append(item[-1])        
                            img_html = "\n".join(img_html)
                            
                            cell_content = img_html + f"\n{text}" if img_html else text
                            html_parts.append(f"<{tag}{attrs}>{cell_content}</{tag}>")
                        html_parts.append("</tr>")
                    html_parts.append("</table>")
                    table_html = "\n".join(html_parts)  
                    sheet_parts = [f"{table_html}"]                              
                    for latex in _extract_drawing_math(_zf, sheet_idx):
                        sheet_parts.append(f"\n$$\n{latex}\n$$")    
                    sheet_content = "\n".join(sheet_parts)
                    sheet_content = clean_text_linebreak(sheet_content)
                    
                    time_elapse = time.perf_counter() - start_time
                    sheet_images = sheet_images if not excel_extract_images else {}
                    res.append({
                        "sheet_idx": sheet_idx+1,
                        "sheet_name": sheet_name,
                        "type": "xlsx",
                        "content": sheet_content,
                        "time_elapse": time_elapse,
                        "images": sheet_images
                    })

            parse_callback.on_finished(**{
                "file": file
            })                            
        except:
            traceback.print_exc()
        finally:
            return res